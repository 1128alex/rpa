from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"])
for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"]
# print(col_B)
for cell in col_B:
    print(cell.value, end=" ")

col_range = ws["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value, end=" ")

row_title = ws[1]
for cell in row_title:
    print(cell.value, end=" ")

row_range = ws[2:6]
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

from openpyxl.utils.cell import coordinate_from_string

row_range = ws[2 : ws.max_row]
for rows in row_range:
    for cell in rows:
        print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")
        print(xy[0], end=" ")  # A
        print(xy[1], end=" ")  # 1
    print()


# all rows
print(tuple(ws.rows))
for row in tuple(ws.rows):
    print(row[2].value)

for row in ws.iter_rows():
    print(row[2].value)

# all columns
print(tuple(ws.columns))
for column in tuple(ws.columns):
    print(column[0].value)

for column in ws.iter_cols():
    print(column[0].value)

# 1번째 줄부터 5번째 줄까지, 2번째 열부터 3번째 열까지
for row in ws.iter_rows(min_row=1, max_row=5, min_col=2, max_col=3):
    print(row[2].value)

wb.save("example.xlsx")
