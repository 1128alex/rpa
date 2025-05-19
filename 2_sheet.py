from openpyxl import Workbook

wb = Workbook()

ws = wb.active
ws.title = "My Sheet"
ws.sheet_properties.tabColor = "1072BA"

ws1 = wb.create_sheet("My Sheet 2")
ws2 = wb.create_sheet("My Sheet 3", 2)

new_ws = wb["NewSheet"]

print(ws.sheetnames)

new_ws["A1"] = "Hello"
target = wb.copy_worksheet(new_ws)
target.title = "Copy of NewSheet"

wb.save("example.xlsx")
