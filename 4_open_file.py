from openpyxl import load_workbook

wb = load_workbook("example.xlsx")
ws = wb.active

for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()
