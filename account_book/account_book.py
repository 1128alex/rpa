import pyautogui
from pyautogui import ImageNotFoundException
from openpyxl import load_workbook

import pyperclip
import re

for w in pyautogui.getWindowsWithTitle("이지원님의 메시지"):
    print(w)

kakaoWindow = pyautogui.getWindowsWithTitle("이지원님의 메시지")[0]
print(kakaoWindow)
if kakaoWindow.isActive == False:
    kakaoWindow.activate()
pyautogui.click(
    kakaoWindow.left + kakaoWindow.width / 2, kakaoWindow.top + kakaoWindow.height / 2
)

pyautogui.hotkey("ctrl", "a")
pyautogui.hotkey("ctrl", "c")

clipboard_text = pyperclip.paste()

lines = clipboard_text.splitlines()
result = []
amounts = []
store_names = []

for i in range(1, len(lines), 2):
    line = lines[i].strip()
    if line:
        tokens = line.split()
        result.append(tokens)

        match = re.search(r"\d[\d,]*", tokens[4])
        amount = int(match.group().replace(",", ""))
        amounts.append(amount)

        store = " ".join(tokens[5:])
        store_names.append(store)

print("Extracted amounts:", amounts)
print("Extracted merchant names:", store_names)

wb = load_workbook("account_book.xlsx")
ws = wb.active
ws.title = "temp"

index = 2
while True:
    if ws.cell(row=index, column=2).value is None:
        break
    index += 1

for store_name in store_names:
    ws.cell(row=index, column=2, value=store_name)
    index += 1
    print("added store name")


index = 2
while True:
    if ws.cell(row=index, column=3).value is None:
        break
    index += 1

for amount in amounts:
    ws.cell(row=index, column=3, value=amount)
    index += 1

for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

clipboard_data = ""

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
    clipboard_data += "\t".join(row_values) + "\n"

pyperclip.copy(clipboard_data.strip())

print("All data copied to clipboard! You can now paste it into Excel Online.")

wb.close()


pyautogui.rightClick(pyautogui.locateOnScreen("chrome.png", confidence=0.8))
pyautogui.write(["up", "up", "up", "enter"], interval=0.25)


while pyautogui.getActiveWindow().title != "새 탭 - Chrome":
    pyautogui.sleep(0.5)

pyautogui.write("excel")
pyautogui.write(["enter"])


def safe_locate_on_screen(*args, **kwargs):
    try:
        return pyautogui.locateOnScreen(*args, **kwargs)
    except ImageNotFoundException:
        return None


print("a")
while pyautogui.getActiveWindow().title != "Excel | Microsoft 365 Copilot - Chrome":
    pyautogui.sleep(1)

pyautogui.moveTo(636, 462)

for i in range(10):
    account_book2 = safe_locate_on_screen("account_book2.png", confidence=0.95)
    if account_book2 is not None:
        print("b")
        pyautogui.click(account_book2)
        break
    else:
        print("c")
        pyautogui.scroll(-300)
        pyautogui.sleep(1)

while pyautogui.getActiveWindow().title != "가계부.xlsx - Chrome":
    pyautogui.sleep(1)
    pyautogui.sleep(2)

pyautogui.click(1700, 900)

pyautogui.hotkey("ctrl", "a")
pyautogui.sleep(3)
pyautogui.hotkey("ctrl", "v")
