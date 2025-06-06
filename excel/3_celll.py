from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "My Sheet"

ws["A1"] = "1"
ws["A2"] = "2"
ws["A3"] = "3"
ws["B1"] = "4"
ws["B2"] = "5"
ws["B3"] = "6"

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value)

print(ws.cell(row=1, column=1).value)
print(ws.cell(row=1, column=2).value)
c = ws.cell(row=1, column=3, value=10)
print(c.value)

index = 1
for x in range(1, 11):
    for y in range(1, 11):
        ws.cell(row=x, column=y, value=index)
        index += 1

wb.save("example.xlsx")
