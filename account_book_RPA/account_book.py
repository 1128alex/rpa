import pyautogui
from pyautogui import ImageNotFoundException
from openpyxl import load_workbook
import sys
import pyperclip
import re

print("Paste the log and press Ctrl+Z:")
clipboard_text = sys.stdin.read()

lines = clipboard_text.splitlines()
dates = []
amounts = []
store_names = []

for i in range(1, len(lines), 2):
    line = lines[i].strip()
    if line:
        tokens = line.split()

        month, day = tokens[2].split("/")
        date = f"{int(month)}월 {int(day)}일"
        dates.append(date)

        match = re.search(r"\d[\d,]*", tokens[4])
        amount = int(match.group().replace(",", ""))
        amounts.append(amount)

        store = " ".join(tokens[5:])
        store_names.append(store)

print("Extracted amounts:", amounts)
print("Extracted merchant names:", store_names)

wb = load_workbook("1account_book.xlsx")
ws = wb.active
ws.title = "temp"

index = 1
while ws.cell(row=index, column=2).value is not None:
    index += 1

last_written_date = None

for date, store, amount in zip(dates, store_names, amounts):
    if date != last_written_date:
        ws.cell(row=index, column=1, value=date)
        last_written_date = date

    ws.cell(row=index, column=2, value=store)
    ws.cell(row=index, column=3, value=amount)

    index += 1

clipboard_data = ""

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
    clipboard_data += "\t".join(row_values) + "\n"

pyperclip.copy(clipboard_data.strip())

wb.save("1account_book.xlsx")
wb.close()

pyautogui.rightClick(pyautogui.locateOnScreen("images/chrome.png", confidence=0.8))
pyautogui.write(["up", "up", "up", "enter"], interval=0.25)


while pyautogui.getActiveWindow().title != "새 탭 - Chrome":
    pyautogui.sleep(0.5)


def safe_locate_on_screen(*args, **kwargs):
    try:
        return pyautogui.locateOnScreen(*args, **kwargs)
    except ImageNotFoundException:
        return None


for i in range(10):
    account_book2 = safe_locate_on_screen("images/account_book2.png", confidence=0.95)
    if account_book2 is not None:
        print("b")
        pyautogui.click(account_book2)
        break
    else:
        print("c")
        pyautogui.scroll(-300)
        pyautogui.sleep(1)

while pyautogui.getActiveWindow().title != "x가계부.xlsx - Chrome":
    pyautogui.sleep(1)
    pyautogui.sleep(2)

pyautogui.click(1700, 900)

pyautogui.sleep(5)
pyautogui.hotkey("ctrl", "a")
pyautogui.sleep(0.2)
pyautogui.hotkey("ctrl", "v")
