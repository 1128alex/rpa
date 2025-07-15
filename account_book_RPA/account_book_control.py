from openpyxl import load_workbook

wb = load_workbook("1account_book.xlsx")
ws = wb.active
ws.title = "temp"

for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

# ws.cell(row=22, column=8, value="SUM(C2:C300)")

wb.save("1account_book.xlsx")
wb.close()
